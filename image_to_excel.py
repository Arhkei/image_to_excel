import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from PIL import Image
from tqdm import tqdm

wb = Workbook()
sheet = wb.active

def rgb_to_hex(rgb):
	return '%02x%02x%02x' % rgb


for file in os.listdir():
	if file.endswith(".jpg") or file.endswith(".jpeg") or file.endswith(".png"):
		filename = file.split(".")[0]
		im = Image.open(file).convert('RGB')
		px = im.load()
		width, height = im.size
		for x in tqdm(range(width)):
			for y in range(height):
				sheet.cell(row=y+1, column=x+1).fill = PatternFill(start_color=rgb_to_hex(px[x,y]), fill_type="solid")

		im.close()
		wb.save(f"{filename}.xlsx")